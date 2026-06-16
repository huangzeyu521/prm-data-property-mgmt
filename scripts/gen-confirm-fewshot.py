#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
把南网 MDAU 产权补录工单(M01 应用域系统 + M02 表级清单)脱敏后生成
确权要素抽取的 few-shot 评测集。每个数据表 = 一个 case:
  input    : 系统上下文(名称/业务域/库类型/登记类型/A-F来源方式标记) + 表中文名/注释
  expected : 工单自带 ground-truth(密级/来源方式/来源主体/G行政监管/H个人隐私/I第三方商密/J第三方协议)
脱敏:丢弃 IP/实例名/schema/系统负责人姓名/联系方式;对保留文本再做电话/IP 兜底擦除。
源数据在仓库外,本脚本可重复执行重生数据集。
用法: python scripts/gen-confirm-fewshot.py [SRC_DIR] [OUT_JSON]
"""
import openpyxl, glob, json, os, re, sys, hashlib

SRC = sys.argv[1] if len(sys.argv) > 1 else r"D:\MyAIProject\nanwangexcel\gydesign\产权"
OUT = sys.argv[2] if len(sys.argv) > 2 else \
    "prm-backend/dpr-confirm-service/src/test/resources/aitool/eval/confirm-fewshot.json"

PHONE = re.compile(r"(?<!\d)\d{11}(?!\d)")
IPV4 = re.compile(r"\b\d{1,3}(?:\.\d{1,3}){3}\b")
SECRECY = {"不涉密", "核心商密", "普通商密", "工作秘密", "敏感信息"}
DATA_ROW_START = 4  # 1=表头 2=图例 3=样例 4+=真实数据
PER_SYSTEM_CAP = 15


def scrub(s):
    if s is None:
        return None
    t = str(s).replace("　", " ").strip()
    t = re.sub(r"\s+", " ", t)
    t = PHONE.sub("[脱敏]", t)
    t = IPV4.sub("[脱敏]", t)
    return t or None


def truthy(v):
    if v is None:
        return False
    t = str(v).strip()
    return t in ("是", "√", "有", "Y", "y", "true", "True")


def norm_source(v):
    """来源方式标准化为 'A 自行生产数据' 形式的首字母大类。"""
    t = scrub(v)
    if not t:
        return None
    m = re.match(r"([A-F])", t)
    return t if m else t


def order_no(fname):
    m = re.match(r"(MDAU-\d{2}-\d{8}-\d{5})", fname)
    return m.group(1) if m else fname[:24]


def read_systems(ws):
    out = []
    for r in ws.iter_rows(min_row=DATA_ROW_START, values_only=True):
        name = scrub(r[0]) if len(r) > 0 else None
        if not name:
            continue
        out.append({
            "name": name,
            "summary": scrub(r[1]) if len(r) > 1 else None,
            "businessDomain": scrub(r[3]) if len(r) > 3 else None,
            "dbType": scrub(r[4]) if len(r) > 4 else None,
            "registerType": scrub(r[18]) if len(r) > 18 else None,
            "sourceFlags": {
                "A自行生产": truthy(r[19]) if len(r) > 19 else False,
                "B公开采集": truthy(r[20]) if len(r) > 20 else False,
                "C公共授权": truthy(r[21]) if len(r) > 21 else False,
                "D公共生产": truthy(r[22]) if len(r) > 22 else False,
                "E交易采购": truthy(r[23]) if len(r) > 23 else False,
                "F其他方式": truthy(r[24]) if len(r) > 24 else False,
            },
        })
    return out


def read_tables(ws):
    out = []
    for r in ws.iter_rows(min_row=DATA_ROW_START, values_only=True):
        cn = scrub(r[3]) if len(r) > 3 else None
        if not cn:
            continue
        out.append({
            "tableNameCn": cn,
            "tableComment": scrub(r[4]) if len(r) > 4 else None,
            "secrecy": scrub(r[5]) if len(r) > 5 else None,
            "sourceType": norm_source(r[6]) if len(r) > 6 else None,
            "sourceSubject": scrub(r[7]) if len(r) > 7 else None,
            "involvesRegulation": truthy(r[10]) if len(r) > 10 else False,
            "regulationSubject": scrub(r[11]) if len(r) > 11 else None,
            "involvesPrivacy": truthy(r[13]) if len(r) > 13 else False,
            "privacyNote": scrub(r[14]) if len(r) > 14 else None,
            "involvesTradeSecret": truthy(r[16]) if len(r) > 16 else False,
            "tradeSecretNote": scrub(r[17]) if len(r) > 17 else None,
            "involvesThirdPartyAgreement": truthy(r[19]) if len(r) > 19 else False,
        })
    return out


def label_key(t):
    # 含表名:仅去除完全重复的整行,保留不同表(评测覆盖更充分)
    return (t["tableNameCn"], t["secrecy"], t["sourceType"], t["involvesRegulation"],
            t["involvesPrivacy"], t["involvesTradeSecret"], t["involvesThirdPartyAgreement"])


cases = []
files = sorted(glob.glob(os.path.join(SRC, "MDAU-*", "MDAU-*", "*_1.xlsx")))
for path in files:
    fname = os.path.basename(path)
    ono = order_no(fname)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as ex:
        print(f"SKIP {fname}: {ex}")
        continue
    m01 = next((w for w in wb.worksheets if w.title.startswith("M01")), None)
    m02 = next((w for w in wb.worksheets if w.title.startswith("M02")), None)
    systems = read_systems(m01) if m01 else []
    sys_ctx = systems[0] if systems else {"name": ono}
    tables = read_tables(m02) if m02 else []
    wb.close()

    seen, kept = set(), 0
    for i, t in enumerate(tables):
        k = label_key(t)
        if k in seen:
            continue
        seen.add(k)
        kept += 1
        if kept > PER_SYSTEM_CAP:
            break
        cid = ono + "#" + hashlib.md5((t["tableNameCn"] + str(i)).encode("utf-8")).hexdigest()[:8]
        cases.append({
            "id": cid,
            "source": ono,
            "system": sys_ctx,
            "input": {
                "tableNameCn": t["tableNameCn"],
                "tableComment": t["tableComment"],
            },
            "expected": {
                "secrecy": t["secrecy"],
                "sourceType": t["sourceType"],
                "sourceSubject": t["sourceSubject"],
                "involvesRegulation": t["involvesRegulation"],
                "regulationSubject": t["regulationSubject"],
                "involvesPrivacy": t["involvesPrivacy"],
                "privacyNote": t["privacyNote"],
                "involvesTradeSecret": t["involvesTradeSecret"],
                "tradeSecretNote": t["tradeSecretNote"],
                "involvesThirdPartyAgreement": t["involvesThirdPartyAgreement"],
            },
        })


def dist(field):
    d = {}
    for c in cases:
        v = c["expected"].get(field)
        d[str(v)] = d.get(str(v), 0) + 1
    return d


dataset = {
    "name": "confirm-element-fewshot",
    "desc": "南网 MDAU 产权补录工单脱敏后的确权要素抽取 few-shot/评测集(每表一例)",
    "sourceWorkOrders": len(files),
    "caseCount": len(cases),
    "labelSchema": {
        "secrecy": sorted(SECRECY),
        "sourceType": "A自行生产/B公开采集/C公共授权/D公共生产/E交易采购/F其他方式",
        "booleans": ["involvesRegulation", "involvesPrivacy", "involvesTradeSecret", "involvesThirdPartyAgreement"],
    },
    "distribution": {
        "secrecy": dist("secrecy"),
        "sourceType": dist("sourceType"),
        "involvesRegulation": dist("involvesRegulation"),
        "involvesPrivacy": dist("involvesPrivacy"),
    },
    "cases": cases,
}

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(dataset, f, ensure_ascii=False, indent=2)
print(f"work orders={len(files)} cases={len(cases)} -> {OUT}")
print("secrecy dist:", dataset["distribution"]["secrecy"])
print("sourceType dist:", dataset["distribution"]["sourceType"])
